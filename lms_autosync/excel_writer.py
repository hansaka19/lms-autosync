import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="12294B")
HEADER_FONT = Font(color="FFFFFF", bold=True)

URGENCY_COLORS = {
    "Overdue": "F8D7DA",
    "Due Soon": "FFF3CD",
    "Upcoming": "D4EDDA",
    "Unknown": "E2E3E5",
}


from datetime import datetime as _dt

CATEGORY_COLORS = {
    "Lab Test": "D6E4F0",
    "Online Quiz": "FCE8D5",
    "Mini Project": "E2D6F0",
    "TMA": "D6F0E0",
    "Other": "E2E3E5",
}


def _parse_moodle_datetime(text):
    """'Wednesday, 3 June 2026, 11:59 PM' වගේ text එකක් sortable datetime
    එකකට convert කරනවා."""
    if not text:
        return None
    try:
        return _dt.strptime(text, "%A, %d %B %Y, %I:%M %p")
    except ValueError:
        return None


def _write_lab_quiz_sheet(wb, lab_quiz_items):
    ws = wb.create_sheet("Lab Tests & Quizzes")
    row_idx = 1
    order = ["Lab Test", "Online Quiz", "Mini Project", "TMA", "Other"]

    for category in order:
        group = [i for i in lab_quiz_items if i.get("category") == category]
        if not group:
            continue
        group.sort(key=lambda i: (
            _parse_moodle_datetime(i.get("due_at_full")) is None,
            _parse_moodle_datetime(i.get("due_at_full")) or _dt.max,
        ))

        ws.cell(row=row_idx, column=1, value=f"{category} ({len(group)})")
        for c in range(1, 6):
            ws.cell(row=row_idx, column=c).fill = PatternFill("solid", fgColor=CATEGORY_COLORS[category])
            ws.cell(row=row_idx, column=c).font = Font(bold=True)
        row_idx += 1

        for c, h in enumerate(["Course", "Item", "Opened", "Due", "Link"], start=1):
            cell = ws.cell(row=row_idx, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row_idx += 1

        for i in group:
            ws.cell(row=row_idx, column=1, value=i["course"])
            ws.cell(row=row_idx, column=2, value=i["name"])
            ws.cell(row=row_idx, column=3, value=i.get("opened_at"))
            ws.cell(row=row_idx, column=4, value=i.get("due_at_full"))
            ws.cell(row=row_idx, column=5, value=i.get("url"))
            row_idx += 1

        row_idx += 1

    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 60)


def _write_sheet(wb, name, headers, rows, color_col=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
        if color_col is not None:
            urgency = row[color_col]
            fill = URGENCY_COLORS.get(urgency)
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=fill)
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 60)


def _write_urgency_sheet(wb, deadlines):
    ws = wb.create_sheet("By Urgency")
    row_idx = 1
    order = ["Overdue", "Due Soon", "Upcoming", "Unknown"]

    for label in order:
        group = [d for d in deadlines if d.get("urgency") == label]
        if not group:
            continue
        group.sort(key=lambda d: (d.get("due_date") is None, d.get("due_date") or ""))

        ws.cell(row=row_idx, column=1, value=f"{label} ({len(group)})")
        for c in range(1, 5):
            ws.cell(row=row_idx, column=c).fill = PatternFill("solid", fgColor=URGENCY_COLORS[label])
            ws.cell(row=row_idx, column=c).font = Font(bold=True)
        row_idx += 1

        for c, h in enumerate(["Due Date", "Course", "Item", "Link"], start=1):
            cell = ws.cell(row=row_idx, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row_idx += 1

        for d in group:
            ws.cell(row=row_idx, column=1, value=d.get("due_date"))
            ws.cell(row=row_idx, column=2, value=d["course"])
            ws.cell(row=row_idx, column=3, value=d["name"])
            ws.cell(row=row_idx, column=4, value=d.get("url"))
            row_idx += 1

        row_idx += 1

    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 60)


def write_workbook(deadlines, resources, announcements=None, lab_quiz_items=None,
                    output_path="output/LMS_Schedule.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    _write_sheet(wb, "Deadlines",
        ["Course", "Item", "Type", "Due Date", "Urgency", "Link"],
        [[d["course"], d["name"], d["type"], d.get("due_date"), d.get("urgency"), d.get("url")]
         for d in deadlines],
        color_col=4)

    _write_urgency_sheet(wb, deadlines)

    if lab_quiz_items:
        _write_lab_quiz_sheet(wb, lab_quiz_items)

    _write_sheet(wb, "Resources",
        ["Course", "Name", "Type", "Link"],
        [[r["course"], r["name"], r["type"], r.get("url")] for r in resources])

    if announcements:
        announcements_sorted = sorted(announcements, key=lambda a: a.get("date") or "", reverse=True)
        _write_sheet(wb, "Announcements",
            ["Course", "Title", "Author", "Date", "Link"],
            [[a["course"], a["title"], a.get("author"), a.get("date"), a.get("url")]
             for a in announcements_sorted])

    summary = wb.create_sheet("Summary", 0)
    overdue = sum(1 for d in deadlines if d.get("urgency") == "Overdue")
    due_soon = sum(1 for d in deadlines if d.get("urgency") == "Due Soon")
    for row in [
        ["Last updated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Total deadlines", len(deadlines)],
        ["Overdue", overdue],
        ["Due Soon", due_soon],
        ["Total resources", len(resources)],
        ["Total announcements", len(announcements or [])],
        ["Lab Tests & Quizzes", len(lab_quiz_items or [])],
    ]:
        summary.append(row)
    for row in summary.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
        row[0].font = Font(bold=True)

    wb.save(output_path)
    print(f"Saved: {output_path}")